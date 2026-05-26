#!/usr/bin/env python3
"""
Extract ketone CGM data from KetoCheck sensor .xlsx exports and output
structured JSON for dashboard HTML generation.

Usage:
    python3 ketone_dashboard_analysis.py

Output: JSON to stdout with per-sensor stats, daily averages, hourly patterns,
distribution buckets, and first/last 24h comparison.

Assumes files exist at the Ketones path below.
"""

import openpyxl
import json
from datetime import datetime
from collections import defaultdict

BASE = '/Users/lyndonarthurson/ORG BOARD/Div 5 (Personal Dev., Health, Correction) /HEALTH/1. Personal data/1. Test Results/Ketones'
FILES = [
    'AAH25B1AA0NX20260521.xlsx',
    'LE2412ZWB620260521.xlsx'
]

def analyse_sensor(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    sensor_id = ws.title

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[1] is not None and row[2] is not None:
            try:
                val = float(row[2])
                rows.append({'no': row[0], 'time': str(row[1]), 'value': val})
            except (ValueError, TypeError):
                pass

    values = [r['value'] for r in rows]
    times = [r['time'] for r in rows]

    # Daily stats
    daily = defaultdict(list)
    for r in rows:
        day = r['time'][:10]
        daily[day].append(r['value'])

    daily_stats = []
    for d in sorted(daily.keys()):
        vals = daily[d]
        daily_stats.append({
            'date': d,
            'count': len(vals),
            'mean': round(sum(vals) / len(vals), 3),
            'min': round(min(vals), 2),
            'max': round(max(vals), 2),
        })

    # Hourly
    hourly = defaultdict(list)
    for r in rows:
        try:
            hr = int(r['time'][11:13])
            hourly[hr].append(r['value'])
        except ValueError:
            pass

    # 4-period breakdown
    periods = {
        'Overnight (00:00-05:59)': list(range(0, 6)),
        'Morning (06:00-11:59)': list(range(6, 12)),
        'Afternoon (12:00-17:59)': list(range(12, 18)),
        'Evening (18:00-23:59)': list(range(18, 24)),
    }
    period_avgs = {}
    for name, hrs in periods.items():
        vals = [v for h in hrs for v in hourly.get(h, [])]
        period_avgs[name] = round(sum(vals) / len(vals), 3) if vals else 0

    # Distribution
    ranges = {
        '<0.5': len([v for v in values if v < 0.5]),
        '0.5-1.0': len([v for v in values if 0.5 <= v < 1.0]),
        '1.0-1.5': len([v for v in values if 1.0 <= v < 1.5]),
        '1.5+': len([v for v in values if v >= 1.5]),
    }

    first_24h = [r['value'] for r in rows[:288]]
    last_24h = [r['value'] for r in rows[-288:]]

    sorted_vals = sorted(values)

    return {
        'sensor_id': sensor_id,
        'total_readings': len(rows),
        'date_start': times[0][:10] if times else None,
        'date_end': times[-1][:10] if times else None,
        'days': len(daily_stats),
        'mean': round(sum(values) / len(values), 3),
        'min': round(min(values), 2),
        'max': round(max(values), 2),
        'median': round(sorted_vals[len(sorted_vals) // 2], 3),
        'daily': daily_stats,
        'ranges': ranges,
        'ranges_pct': {k: round(v / len(values) * 100, 1) for k, v in ranges.items()},
        'first_24h_mean': round(sum(first_24h) / len(first_24h), 3) if first_24h else None,
        'last_24h_mean': round(sum(last_24h) / len(last_24h), 3) if last_24h else None,
        'period_avgs': period_avgs,
    }


def main():
    results = {}
    for fname in FILES:
        fp = f'{BASE}/{fname}'
        print(f'Analysing: {fname}', file=open('/dev/stderr', 'w'))
        results[fname] = analyse_sensor(fp)
    print(json.dumps(results, indent=2))


if __name__ == '__main__':
    main()
